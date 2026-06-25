# Genera el workbook de 7 hojas a partir de data = {key: DataFrame(fecha, vc, patrimonio)}
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORDER_DEFAULT=["01_opportunity","02_rentafija1","03_growth","04_vivienda01","05_bluewhale","06_commercial"]
DISP_DEFAULT={"01_opportunity":"FI Abierto Hencorp Opportunity","02_rentafija1":"FIC Renta Fija I",
      "03_growth":"FIC Inmobiliario Hencorp Growth","04_vivienda01":"FIC Desarrollo Inmob. Hencorp Vivienda 01",
      "05_bluewhale":"FIC Inmobiliario Hencorp Blue Whale","06_commercial":"FIC Inmobiliario Hencorp Commercial Properties"}
TIPO_DEFAULT={"01_opportunity":"Abierto","02_rentafija1":"Cerrado","03_growth":"Cerrado","04_vivienda01":"Cerrado","05_bluewhale":"Cerrado","06_commercial":"Cerrado"}
ABBR={1:"ene",2:"feb",3:"mar",4:"abr",5:"may",6:"jun",7:"jul",8:"ago",9:"sep",10:"oct",11:"nov",12:"dic"}

def build(data, out_path, corte_label="datos actualizados automáticamente",
          order=None, disp=None, tipo=None, sub_prefix=None,
          titulo_resumen=None, fuente_nota=None, notas_fondos=None):
    ORDER = order if order is not None else ORDER_DEFAULT
    DISP  = disp  if disp  is not None else DISP_DEFAULT
    TIPO  = tipo  if tipo  is not None else TIPO_DEFAULT
    if sub_prefix is None:
        sub_prefix = "Grupo Hencorp \u2014 Gestora de Fondos | Fuente: hencorpgestora.com (Comportamiento Historico)"
    if titulo_resumen is None:
        titulo_resumen = f"Resumen \u2014 {len(ORDER)} Fondos de Inversi\u00f3n Hencorp"
    if fuente_nota is None:
        fuente_nota = "Fuente: Hencorp Gestora de Fondos de Inversi\u00f3n, S.A. \u2014 secci\u00f3n \u0027Informe Valor Cuota\u0027 (PDF \u0027Comportamiento Hist\u00f3rico del Fondo\u0027) en hencorpgestora.com. Archivo regenerado autom\u00e1ticamente."
    F={}
    for k in ORDER:
        df=data[k].copy()
        me_vc=df.set_index("fecha")["vc"].resample("ME").last().dropna()
        me_pat=df.set_index("fecha")["patrimonio"].resample("ME").last()
        F[k]=dict(vc={d.strftime("%Y-%m"):float(v) for d,v in me_vc.items()},
                  pat={d.strftime("%Y-%m"):(None if pd.isna(v) else float(v)) for d,v in me_pat.items()},
                  base=float(df["vc"].iloc[0]), inc=me_vc.index[0].strftime("%Y-%m"),
                  first_date=df["fecha"].min().to_pydatetime(), last_date=df["fecha"].max().to_pydatetime(),
                  nmonths=len(me_vc))
    start=min(pd.Period(F[k]["inc"],freq="M") for k in ORDER)
    end=max(pd.Period(F[k]["vc"] and max(F[k]["vc"].keys()),freq="M") for k in ORDER)
    allm=pd.period_range(start,end,freq="M")
    months=[(f"{ABBR[p.month]}-{p.year}"+("*" if (p.year==end.year and p.month==end.month) else ""), f"{p.year:04d}-{p.month:02d}") for p in allm]
    ym2col={ym:get_column_letter(2+i) for i,(lab,ym) in enumerate(months)}
    last_ym=f"{end.year:04d}-{end.month:02d}"
    years=list(range(min(int(F[k]['inc'][:4]) for k in ORDER), end.year+1))

    ARIAL=lambda **kw: Font(name="Arial", **kw)
    hdr=PatternFill("solid",fgColor="1F3864"); fund_fill=PatternFill("solid",fgColor="F2F2F2")
    white=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    thin=Side(style="thin",color="BFBFBF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
    BLUE="0000FF"; BLACK="000000"; GREEN="006100"
    SH_VC="Valor cuota (fin de mes)"; SH_PAT="Patrimonio (fin de mes)"; SH_MEN="Rendimiento mensual"
    SH_DIC="Cierre 31-dic (mensual)"; SH_ANU="Rendimiento anual"; SH_RES="Resumen"; SH_NOT="Notas y metodologia"
    PCT="0.00%"; VCFMT="#,##0.000000"; USD="$#,##0"; DT="dd/mm/yyyy"
    SUB=f"{sub_prefix} | {corte_label}"
    wb=Workbook()

    def title_block(ws,title,sub,ncols):
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=ncols)
        c=ws.cell(1,1,title); c.font=ARIAL(bold=True,size=13,color="1F3864"); c.alignment=left
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=ncols)
        c=ws.cell(2,1,sub); c.font=ARIAL(italic=True,size=9,color="595959"); c.alignment=left
        ws.row_dimensions[1].height=20

    def set_val(ws,r,c,f,k,ck,kind):
        cell=ws.cell(r,c); col=get_column_letter(c); base_ref=f"'{SH_RES}'!$D${4+f}"
        if kind=="vc":
            v=F[k]["vc"].get(ck)
            if v is not None: cell.value=v; cell.number_format=VCFMT; cell.font=ARIAL(color=BLUE,size=9)
        elif kind=="pat":
            v=F[k]["pat"].get(ck)
            if v is not None: cell.value=v; cell.number_format=USD; cell.font=ARIAL(color=BLUE,size=9)
        elif kind=="ret":
            if ck in F[k]["vc"]:
                vccell=f"'{SH_VC}'!{col}{r}"
                if ck==F[k]["inc"]:
                    cell.value=f'=IF({vccell}="","",{vccell}/{base_ref}-1)'
                else:
                    prev=get_column_letter(c-1); prevcell=f"'{SH_VC}'!{prev}{r}"
                    cell.value=f'=IFERROR(IF(OR({vccell}="",{prevcell}=""),"",{vccell}/{prevcell}-1),"")'
                cell.number_format=PCT; cell.font=ARIAL(color=BLACK,size=9)

    def matrix_sheet(name,col_labels,col_keys,kind):
        ws=wb.create_sheet(name); ncols=1+len(col_labels); title_block(ws,name,SUB,ncols)
        h=ws.cell(3,1,"Fondo"); h.font=white; h.fill=hdr; h.alignment=left; h.border=border
        for j,lab in enumerate(col_labels):
            c=ws.cell(3,2+j,lab); c.font=white; c.fill=hdr; c.alignment=center; c.border=border
        for f,k in enumerate(ORDER):
            r=4+f; cc=ws.cell(r,1,DISP[k]); cc.font=ARIAL(bold=True,size=9); cc.fill=fund_fill; cc.alignment=left; cc.border=border
            for j,ck in enumerate(col_keys):
                ws.cell(r,2+j).border=border; ws.cell(r,2+j).alignment=center; set_val(ws,r,2+j,f,k,ck,kind)
        ws.freeze_panes="B4"; ws.column_dimensions["A"].width=42
        for j in range(len(col_labels)): ws.column_dimensions[get_column_letter(2+j)].width=12 if kind in("vc","ret") else 16

    mlabels=[m[0] for m in months]; mkeys=[m[1] for m in months]
    matrix_sheet(SH_VC,mlabels,mkeys,"vc"); matrix_sheet(SH_PAT,mlabels,mkeys,"pat"); matrix_sheet(SH_MEN,mlabels,mkeys,"ret")

    def simple_sheet(name,sub):
        ws=wb.create_sheet(name); ncols=1+len(years); title_block(ws,name,sub,ncols)
        h=ws.cell(3,1,"Fondo"); h.font=white; h.fill=hdr; h.border=border; h.alignment=left
        for j,y in enumerate(years):
            c=ws.cell(3,2+j,f"{end.year} (YTD)" if y==end.year else str(y)); c.font=white; c.fill=hdr; c.alignment=center; c.border=border
        for f,k in enumerate(ORDER):
            r=4+f; cc=ws.cell(r,1,DISP[k]); cc.font=ARIAL(bold=True,size=9); cc.fill=fund_fill; cc.border=border; cc.alignment=left
            for j in range(len(years)): ws.cell(r,2+j).border=border; ws.cell(r,2+j).alignment=center
        ws.freeze_panes="B4"; ws.column_dimensions["A"].width=42
        for j in range(len(years)): ws.column_dimensions[get_column_letter(2+j)].width=13
        return ws

    wsd=simple_sheet(SH_DIC,"Rendimiento del mes de diciembre de cada año (mismo indicador que la hoja mensual). El año en curso = ultimo mes disponible (parcial)")
    for f,k in enumerate(ORDER):
        r=4+f
        for j,y in enumerate(years):
            tgt=last_ym if y==end.year else f"{y}-12"
            if tgt in F[k]["vc"]:
                c=wsd.cell(r,2+j); c.value=f"='{SH_MEN}'!{ym2col[tgt]}{r}"; c.number_format=PCT; c.font=ARIAL(size=9,color=BLACK)

    wsa=simple_sheet(SH_ANU,"Rendimiento ANUAL (ene–dic) desde la cuota. Año de lanzamiento = desde la base; el año en curso = YTD (parcial)")
    for f,k in enumerate(ORDER):
        r=4+f; base_ref=f"'{SH_RES}'!$D${r}"
        for j,y in enumerate(years):
            end_ym=last_ym if y==end.year else f"{y}-12"
            if end_ym not in F[k]["vc"]: continue
            endcell=f"'{SH_VC}'!{ym2col[end_ym]}{r}"; prior=f"{y-1}-12"
            prevref=f"'{SH_VC}'!{ym2col[prior]}{r}" if prior in F[k]["vc"] else base_ref
            c=wsa.cell(r,2+j); c.value=f'=IFERROR({endcell}/{prevref}-1,"")'; c.number_format=PCT; c.font=ARIAL(size=9,color=BLACK)

    ws=wb.create_sheet(SH_RES)
    cols=["Fondo","Tipo","Inicio","VC inicial (base)","VC actual","Fecha dato","Patrimonio actual (US$)","Meses","Rend. acum. desde inicio","Rend. anualizado (CAGR)"]
    title_block(ws,titulo_resumen,SUB,len(cols))
    for j,h in enumerate(cols):
        c=ws.cell(3,1+j,h); c.font=white; c.fill=hdr; c.border=border
        c.alignment=Alignment(horizontal=("left" if j==0 else "center"),vertical="center",wrap_text=True)
    ws.row_dimensions[3].height=30
    for f,k in enumerate(ORDER):
        r=4+f
        ws.cell(r,1,DISP[k]).font=ARIAL(bold=True,size=9); ws.cell(r,2,TIPO[k]).font=ARIAL(size=9)
        ws.cell(r,3,F[k]["first_date"]).number_format=DT; ws.cell(r,3).font=ARIAL(size=9)
        b=ws.cell(r,4,F[k]["base"]); b.number_format=("0.0000" if F[k]["base"]<10 else "#,##0.0000"); b.font=ARIAL(color=BLUE,size=9)
        e=ws.cell(r,5,f"='{SH_VC}'!{ym2col[last_ym]}{r}"); e.number_format=VCFMT; e.font=ARIAL(color=GREEN,size=9)
        ws.cell(r,6,F[k]["last_date"]).number_format=DT; ws.cell(r,6).font=ARIAL(size=9)
        g=ws.cell(r,7,f"='{SH_PAT}'!{ym2col[last_ym]}{r}"); g.number_format=USD; g.font=ARIAL(color=GREEN,size=9)
        ws.cell(r,8,F[k]["nmonths"]).font=ARIAL(size=9)
        i9=ws.cell(r,9,f"=E{r}/D{r}-1"); i9.number_format=PCT; i9.font=ARIAL(size=9)
        j10=ws.cell(r,10,f'=IFERROR((E{r}/D{r})^(365/(F{r}-C{r}))-1,"")'); j10.number_format=PCT; j10.font=ARIAL(size=9)
        for cc in range(1,len(cols)+1):
            ws.cell(r,cc).border=border; ws.cell(r,cc).alignment=(left if cc==1 else center)
    ws.column_dimensions["A"].width=42
    for j,w in zip(range(2,11),[10,12,15,14,13,20,8,18,18]): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="B4"

    wsn=wb.create_sheet(SH_NOT)
    notas=[("Notas y metodología",16,"1F3864",True),("",10,"000000",False),
     (fuente_nota,10,"000000",False),
     (f"{corte_label}.",10,"000000",False),("",10,"000000",False),
     ("DEFINICIÓN DE 'RENDIMIENTO MENSUAL':",11,"1F3864",True),
     ("Rendimiento del mes = (Valor cuota al cierre del mes ÷ Valor cuota al cierre del mes anterior) − 1. Es el rendimiento real de la cuota en el mes.",10,"000000",False),
     ("El primer mes de cada fondo se mide desde el valor cuota inicial (base) hasta el cierre de ese mes (puede ser parcial). El año en curso es PARCIAL.",10,"C00000",True),("",10,"000000",False),
     *(notas_fondos if notas_fondos is not None else [
       ("ADVERTENCIA (fondos cerrados):",11,"C00000",True),
       ("• Renta Fija I DISTRIBUYE en mayo y noviembre: la cuota baja esos meses, por lo que su rendimiento medido SOLO desde la cuota SUBESTIMA el total del partícipe (lo distribuido se paga en efectivo).",10,"000000",False),
       ("• Growth y Vivienda 01 registran REVALUACIONES inmobiliarias (saltos grandes) y bajas por distribución/markdown. Commercial Properties tiene historia mínima (inició may-2026).",10,"000000",False),
     ]),("",10,"000000",False),
     ("Códigos de color: azul = dato de origen extraído; verde = enlace a otra hoja; negro = fórmula.",9,"595959",False)]
    for i,(txt,sz,color,bold) in enumerate(notas,start=1):
        c=wsn.cell(i,1,txt); c.font=ARIAL(size=sz,color=color,bold=bold); c.alignment=Alignment(wrap_text=True,vertical="top")
    wsn.column_dimensions["A"].width=140

    # ===================== HOJA "DIARIO" =====================
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    PALETTE=["1F3864","2E8B57","C0504D","E0A800","7E57C2","0E7C86"]
    f_b=Font(name="Arial",color=BLUE,size=9); f_k=Font(name="Arial",color=BLACK,size=9)
    SHORT={k:DISP[k].replace("FIC ","").replace("FI ","") for k in ORDER}
    mlabels=[m[0] for m in months]; mkeys=[m[1] for m in months]
    daily={}
    for k in ORDER:
        dd=data[k].copy().sort_values("fecha")
        rr=dd["vc"].pct_change(); rr.iloc[0]=dd["vc"].iloc[0]/F[k]["base"]-1
        dd["ret"]=rr; daily[k]=dd.set_index("fecha")
    alldates=sorted(set().union(*[set(daily[k].index) for k in ORDER]), reverse=True)
    wsD=wb.create_sheet("Diario"); ncolsD=1+3*len(ORDER)
    title_block(wsD,f"Detalle diario — {len(ORDER)} fondos de inversión", SUB+" | una fila por día · lo más reciente arriba · se actualiza a diario", ncolsD)
    wsD.merge_cells(start_row=3,start_column=1,end_row=4,end_column=1)
    hc=wsD.cell(3,1,"Fecha"); hc.font=white; hc.fill=hdr; hc.alignment=center; hc.border=border
    mets=["Valor cuota","Rend. diario","Patrimonio"]; col=2
    subfill=PatternFill("solid",fgColor="2E4A7D"); subfont=Font(name="Arial",bold=True,color="FFFFFF",size=8)
    for k in ORDER:
        wsD.merge_cells(start_row=3,start_column=col,end_row=3,end_column=col+2)
        gc=wsD.cell(3,col,SHORT[k]); gc.font=white; gc.fill=hdr; gc.alignment=center; gc.border=border
        for mi,m in enumerate(mets):
            c=wsD.cell(4,col+mi,m); c.font=subfont; c.fill=subfill; c.alignment=center; c.border=border
        col+=3
    for ri,dt in enumerate(alldates, start=5):
        dc=wsD.cell(ri,1, dt.to_pydatetime()); dc.number_format=DT; dc.font=f_k; dc.alignment=center
        col=2
        for k in ORDER:
            if dt in daily[k].index:
                row=daily[k].loc[dt]
                a=wsD.cell(ri,col, round(float(row["vc"]),8)); a.number_format=VCFMT; a.font=f_b; a.alignment=center
                rv=row["ret"]; bcell=wsD.cell(ri,col+1, None if pd.isna(rv) else round(float(rv),6)); bcell.number_format=PCT; bcell.font=f_k; bcell.alignment=center
                pv=row["patrimonio"]; pcell=wsD.cell(ri,col+2, None if pd.isna(pv) else round(float(pv),2)); pcell.number_format=USD; pcell.font=f_b; pcell.alignment=center
            col+=3
    wsD.freeze_panes="B5"; wsD.column_dimensions["A"].width=12
    for cc in range(2,ncolsD+1): wsD.column_dimensions[get_column_letter(cc)].width=13

    # ============== DATOS DE APOYO (oculta) + 18 GRÁFICAS POR FONDO ==============
    shG=wb.create_sheet("Datos_graficas"); shG.sheet_state="hidden"
    def annual_pat(k,y):
        end_ym=last_ym if y==end.year else f"{y}-12"
        return F[k]["pat"].get(end_ym)
    def monthly_ret(k,ym):
        if ym not in F[k]["vc"]: return None
        i=mkeys.index(ym)
        if ym==F[k]["inc"]: return F[k]["vc"][ym]/F[k]["base"]-1
        pv=mkeys[i-1] if i>0 else None
        return F[k]["vc"][ym]/F[k]["vc"][pv]-1 if (pv and pv in F[k]["vc"]) else None
    # Patrimonio mensual (filas 2-7, encabezado fila 1)
    shG.cell(1,1,"Fondo")
    for j,lab in enumerate(mlabels): shG.cell(1,2+j,lab)
    for f,k in enumerate(ORDER):
        shG.cell(2+f,1,SHORT[k])
        for j,ym in enumerate(mkeys):
            p=F[k]["pat"].get(ym)
            if p is not None: shG.cell(2+f,2+j, round(p,2))
    nmP=1+len(months)
    # Patrimonio anual (filas 11-16, encabezado fila 10)
    shG.cell(10,1,"Fondo")
    for j,y in enumerate(years): shG.cell(10,2+j,str(y))
    for f,k in enumerate(ORDER):
        shG.cell(11+f,1,SHORT[k])
        for j,y in enumerate(years):
            pv=annual_pat(k,y)
            if pv is not None: shG.cell(11+f,2+j, round(pv,2))
    nmAy=1+len(years)
    # Rendimiento mensual (filas 20-25, encabezado fila 19)
    shG.cell(19,1,"Fondo")
    for j,lab in enumerate(mlabels): shG.cell(19,2+j,lab)
    for f,k in enumerate(ORDER):
        shG.cell(20+f,1,SHORT[k])
        for j,ym in enumerate(mkeys):
            rv=monthly_ret(k,ym)
            if rv is not None: shG.cell(20+f,2+j, round(rv,6))

    NICK={}
    for k in ORDER:
        d=DISP[k]
        NICK[k]=("Opportunity" if "Opportunity" in d else "Renta Fija I" if "Renta Fija" in d
                 else "Growth" if "Growth" in d else "Vivienda 01" if "Vivienda" in d
                 else "Blue Whale" if "Blue Whale" in d else "Commercial" if "Commercial" in d else SHORT[k])
    skip=max(1, round(len(months)/8))

    graf=wb.create_sheet("Gráficas")
    title_block(graf,"Gráficas por fondo — patrimonio y rendimiento", SUB+" | una gráfica por fondo · cada una con su propio eje · se actualizan solas al crecer meses/días", 22)

    def make_bar(title, datarow, headerrow, lastcol, clr, numfmt, is_monthly):
        ch=BarChart(); ch.type="col"; ch.grouping="clustered"; ch.title=title; ch.height=7; ch.width=10; ch.gapWidth=40
        ch.add_data(Reference(shG,min_col=2,max_col=lastcol,min_row=datarow,max_row=datarow), from_rows=True, titles_from_data=False)
        ch.set_categories(Reference(shG,min_col=2,max_col=lastcol,min_row=headerrow,max_row=headerrow))
        ch.legend=None
        ch.y_axis.numFmt=numfmt; ch.y_axis.delete=False; ch.x_axis.delete=False; ch.x_axis.majorGridlines=None
        if is_monthly:
            ch.x_axis.tickLblSkip=skip; ch.x_axis.tickMarkSkip=skip
        ch.series[0].graphicalProperties=GraphicalProperties(solidFill=clr)
        return ch

    for f,k in enumerate(ORDER):
        clr=PALETTE[f % len(PALETTE)]; br=3+f*17; cr=4+f*17
        for cc in range(1,23): graf.cell(br,cc).fill=hdr
        graf.merge_cells(start_row=br,start_column=1,end_row=br,end_column=22)
        b=graf.cell(br,1,"   "+DISP[k]); b.font=Font(name="Arial",bold=True,color="FFFFFF",size=11); b.alignment=Alignment(horizontal="left",vertical="center")
        graf.row_dimensions[br].height=20
        graf.add_chart(make_bar(f"Patrimonio mensual — {NICK[k]} (US$)", 2+f, 1, nmP, clr, '"$"#,##0,,"M"', True),  f"A{cr}")
        graf.add_chart(make_bar(f"Patrimonio anual — {NICK[k]} (US$)",  11+f, 10, nmAy, clr, '"$"#,##0,,"M"', False), f"H{cr}")
        graf.add_chart(make_bar(f"Rendimiento mensual — {NICK[k]}",     20+f, 19, nmP, clr, '0.0%', True),           f"O{cr}")

    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    desired=[SH_RES,"Gráficas",SH_MEN,SH_DIC,SH_ANU,"Diario",SH_VC,SH_PAT,SH_NOT,"Datos_graficas"]
    wb._sheets.sort(key=lambda s: desired.index(s.title))
    wb.save(out_path)
